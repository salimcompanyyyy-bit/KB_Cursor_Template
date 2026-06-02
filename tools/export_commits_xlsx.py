"""Выгрузка коммитов за день/период в Excel.

Использование:
    py -3 tools/export_commits_xlsx.py
    py -3 tools/export_commits_xlsx.py --since 2026-05-25 --until 2026-05-25
    py -3 tools/export_commits_xlsx.py --output reports/report.xlsx

По умолчанию пишет в ``reports/Коммиты_YYYY_MM_DD.xlsx`` (папка `reports/`
исключена из git через ``.gitignore``). Колонки: # · Время · Хеш · Версия ·
Тип · Описание · Изменения · Автор. Тип подсвечивается заливкой (функция /
исправление / контекст и т.д.); «Изменения» — нумерованный список из тела
коммита (1–3 пункта).
"""

from __future__ import annotations

import argparse
import re
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

COMMIT_RE = re.compile(
    r"^v(?P<version>\d+\.\d+\.\d+):\s*(?P<type>\S+)\s*-\s*(?P<desc>.+)$"
)

TYPE_FILL = {
    "функция": "C6EFCE",
    "исправление": "FFEB9C",
    "контекст": "BDD7EE",
    "добавление": "D9E1F2",
    "рефакторинг": "F4B084",
    "база": "B4A7D6",
    "безопасность": "FFC7CE",
    "производительность": "9FC5E8",
}

# Разделители для безопасного парсинга многострочного %b.
# RS (record separator) — между полями коммита; US (unit separator) — между коммитами.
_RS = "\x1e"
_US = "\x1f"


def fetch_commits(since: str, until: str) -> list[dict]:
    pretty = f"%h{_RS}%ai{_RS}%s{_RS}%b{_RS}%an{_US}"
    cmd = [
        "git",
        "log",
        f"--since={since} 00:00:00",
        f"--until={until} 23:59:59",
        f"--pretty=format:{pretty}",
        "--reverse",
    ]
    out = subprocess.check_output(cmd, text=True, encoding="utf-8")
    rows: list[dict] = []
    for rec in out.split(_US):
        rec = rec.strip("\n\r ")
        if not rec:
            continue
        parts = rec.split(_RS)
        if len(parts) < 5:
            continue
        h, dt, subj, body, author = parts[0], parts[1], parts[2], parts[3], parts[4]
        m = COMMIT_RE.match(subj.strip())
        version = m.group("version") if m else ""
        ctype = m.group("type") if m else ""
        desc = m.group("desc") if m else subj
        rows.append(
            {
                "hash": h.strip(),
                "time": dt[11:16],
                "datetime": dt.strip(),
                "version": version,
                "type": ctype,
                "desc": desc.strip(),
                "body": body.strip(),
                "author": author.strip(),
            }
        )
    return rows


def _estimate_body_height(body: str, col_width: int) -> float:
    """Грубая оценка высоты строки под `body` при заданной ширине колонки."""
    if not body:
        return 18.0
    lines = body.splitlines()
    visual_lines = 0
    chars_per_line = max(10, int(col_width * 1.05))
    for ln in lines:
        if not ln:
            visual_lines += 1
            continue
        visual_lines += max(1, (len(ln) + chars_per_line - 1) // chars_per_line)
    return float(max(20, min(220, visual_lines * 15 + 4)))


def build_xlsx(rows: list[dict], output_path: Path, title: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Коммиты"

    headers = ["#", "Время", "Хеш", "Версия", "Тип", "Описание", "Изменения", "Автор"]
    widths = [4, 8, 10, 9, 16, 45, 70, 18]
    body_col_idx = headers.index("Изменения") + 1
    body_col_width = widths[body_col_idx - 1]

    bottom_thick = Side(style="thick", color="000000")
    bottom_thin = Side(style="hair", color="BFBFBF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, name in enumerate(headers, start=1):
        c = ws.cell(2, col, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = align_center
        c.border = Border(bottom=bottom_thick)

    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    for i, r in enumerate(rows, start=1):
        row_idx = 2 + i
        ws.cell(row_idx, 1, i).alignment = align_center
        ws.cell(row_idx, 2, r["time"]).alignment = align_center
        ws.cell(row_idx, 3, r["hash"]).alignment = align_center
        ws.cell(row_idx, 4, r["version"]).alignment = align_center
        c_type = ws.cell(row_idx, 5, r["type"])
        c_type.alignment = align_center
        fill_color = TYPE_FILL.get(r["type"])
        if fill_color:
            c_type.fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row_idx, 6, r["desc"]).alignment = align_left
        ws.cell(row_idx, 7, r["body"]).alignment = align_top_left
        ws.cell(row_idx, 8, r["author"]).alignment = align_center
        for col in range(1, len(headers) + 1):
            ws.cell(row_idx, col).border = Border(bottom=bottom_thin)
        ws.row_dimensions[row_idx].height = _estimate_body_height(
            r["body"], body_col_width
        )

    summary_row = 2 + len(rows) + 1
    type_counts: dict[str, int] = {}
    for r in rows:
        type_counts[r["type"]] = type_counts.get(r["type"], 0) + 1
    parts = [f"Всего: {len(rows)}"] + [
        f"{t}: {n}" for t, n in sorted(type_counts.items()) if t
    ]
    summary_cell = ws.cell(summary_row, 1, "    ·    ".join(parts))
    summary_cell.font = Font(bold=True)
    ws.merge_cells(
        start_row=summary_row,
        start_column=1,
        end_row=summary_row,
        end_column=len(headers),
    )
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_cell.border = Border(bottom=bottom_thick)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22

    wb.save(output_path)
    return output_path


def main() -> None:
    today = datetime.now().strftime("%Y-%m-%d")
    p = argparse.ArgumentParser()
    p.add_argument("--since", default=today, help="дата с (YYYY-MM-DD), по умолчанию сегодня")
    p.add_argument("--until", default=today, help="дата по (YYYY-MM-DD), по умолчанию сегодня")
    p.add_argument(
        "--output",
        default=None,
        help="путь к xlsx; по умолчанию reports/Коммиты_YYYY_MM_DD.xlsx",
    )
    args = p.parse_args()

    rows = fetch_commits(args.since, args.until)
    if args.since == args.until:
        title = f"Коммиты за {args.since}"
    else:
        title = f"Коммиты {args.since} — {args.until}"

    if args.output:
        out_path = Path(args.output)
    else:
        date_safe = args.since.replace("-", "_")
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        out_path = reports_dir / f"Коммиты_{date_safe}.xlsx"

    build_xlsx(rows, out_path, title)
    print(f"Сохранено: {out_path.resolve()} ({len(rows)} коммитов)")


if __name__ == "__main__":
    main()
